import sys
import json
import re
import openpyxl

# Exact row mapping matching template sheet "02_Assumptions"
STANDARD_ROW_MAP = {
    6: "beginning_cooperatives",
    7: "new_coops_acquired",
    8: "monthly_churn_rate",
    9: "avg_members_per_coop",
    10: "subscription_paying_frac",
    13: "setup_fee",
    14: "paid_implementation_coops",
    15: "monthly_subscription_fee",
    16: "ios_addon_monthly_fee",
    17: "ios_adoption_frac",
    18: "white_label_projects",
    19: "white_label_fee_per_project",
    20: "ppob_active_coops_frac",
    21: "ppob_tx_per_coop_month",
    22: "avg_ppob_fee_per_tx",
    23: "academy_participants_frac",
    24: "academy_avg_price_per_participant",
    25: "offline_trainings_per_month",
    26: "offline_training_fee_per_coop",
    27: "enterprise_api_revenue",
    30: "cloud_cost_per_coop_month",
    31: "implementation_cost_per_coop",
    32: "support_cost_per_coop_month",
    33: "payment_api_var_cost_frac",
    34: "other_cost_of_revenue_frac",
    37: "seed_investment",
    38: "pre_money_valuation",
    39: "exit_revenue_multiple_conservative",
    40: "exit_revenue_multiple_base",
    41: "exit_revenue_multiple_optimistic",
}

# Row mapping for sheet "06_HR_Planning"
HR_ROW_MAP = {
    5: "hr_engineering_fte",
    6: "hr_sales_fte",
    7: "hr_marketing_fte",
    8: "hr_support_fte",
    9: "hr_finance_admin_fte",
    10: "hr_management_fte",
    12: "hr_avg_salary_monthly",
    13: "payroll_cost",
}

# Row mapping for sheet "07_OPEX"
OPEX_ROW_MAP = {
    5: "payroll_cost",
    6: "sales_marketing_spend",
    7: "office_utilities_internet",
    8: "software_tools_subscriptions",
    9: "legal_accounting_compliance",
    10: "travel_events",
    11: "recruitment_training",
    12: "other_ga",
}

FRAC_KEYS = {
    "monthly_churn_rate",
    "subscription_paying_frac",
    "ios_adoption_frac",
    "ppob_active_coops_frac",
    "academy_participants_frac",
    "payment_api_var_cost_frac",
    "other_cost_of_revenue_frac",
}

KEY_ALIASES = {
    "beginning_cooperatives": ["beginning_cooperatives", "beginning active cooperatives", "koperasi awal"],
    "new_coops_acquired": ["new_coops_acquired", "new cooperatives acquired", "koperasi baru"],
    "monthly_churn_rate": ["monthly_churn_rate", "monthly churn rate", "tingkat churn"],
    "avg_members_per_coop": ["avg_members_per_coop", "average members", "anggota per koperasi"],
    "subscription_paying_frac": ["subscription_paying_frac", "subscription-paying %", "pelanggan aktif membayar"],
    "setup_fee": ["setup_fee", "setup fee", "biaya setup"],
    "paid_implementation_coops": ["paid_implementation_coops", "paid implementation coops"],
    "monthly_subscription_fee": ["monthly_subscription_fee", "monthly subscription fee", "biaya langganan bulanan"],
    "ios_addon_monthly_fee": ["ios_addon_monthly_fee", "ios add-on monthly fee"],
    "ios_adoption_frac": ["ios_adoption_frac", "ios adoption %"],
    "white_label_projects": ["white_label_projects", "white label projects"],
    "white_label_fee_per_project": ["white_label_fee_per_project", "white label fee"],
    "ppob_active_coops_frac": ["ppob_active_coops_frac", "ppob active cooperatives %"],
    "ppob_tx_per_coop_month": ["ppob_tx_per_coop_month", "ppob transactions"],
    "avg_ppob_fee_per_tx": ["avg_ppob_fee_per_tx", "average ppob fee"],
    "academy_participants_frac": ["academy_participants_frac", "academy participants %"],
    "academy_avg_price_per_participant": ["academy_avg_price_per_participant", "academy average price"],
    "offline_trainings_per_month": ["offline_trainings_per_month", "offline trainings / month"],
    "offline_training_fee_per_coop": ["offline_training_fee_per_coop", "offline training fee"],
    "enterprise_api_revenue": ["enterprise_api_revenue", "enterprise", "banking", "api revenue"],
    "cloud_cost_per_coop_month": ["cloud_cost_per_coop_month", "cloud cost"],
    "implementation_cost_per_coop": ["implementation_cost_per_coop", "implementation cost"],
    "support_cost_per_coop_month": ["support_cost_per_coop_month", "support cost"],
    "payment_api_var_cost_frac": ["payment_api_var_cost_frac", "payment / api variable cost"],
    "other_cost_of_revenue_frac": ["other_cost_of_revenue_frac", "other cost of revenue"],
    "hr_engineering_fte": ["engineering / product fte", "engineering fte", "hr_engineering_fte"],
    "hr_sales_fte": ["sales & partnership fte", "sales fte", "hr_sales_fte"],
    "hr_marketing_fte": ["marketing fte", "hr_marketing_fte"],
    "hr_support_fte": ["customer success / support fte", "support fte", "hr_support_fte"],
    "hr_finance_admin_fte": ["finance / hr / admin fte", "finance fte", "hr_finance_admin_fte"],
    "hr_management_fte": ["management / leadership fte", "leadership fte", "hr_management_fte"],
    "hr_avg_salary_monthly": ["average salary / month", "gaji bulanan", "hr_avg_salary_monthly"],
    "payroll_cost": ["payroll", "total payroll", "beban gaji", "payroll_cost"],
    "sales_marketing_spend": ["sales & marketing spend", "penjualan & pemasaran", "sales_marketing_spend"],
    "office_utilities_internet": ["office, utilities", "sewa kantor", "office_utilities_internet"],
    "software_tools_subscriptions": ["software tools", "langganan software", "software_tools_subscriptions"],
    "legal_accounting_compliance": ["legal, accounting", "hukum & akuntansi", "legal_accounting_compliance"],
    "travel_events": ["travel & events", "perjalanan dinas", "travel_events"],
    "recruitment_training": ["recruitment & training", "rekrutmen", "recruitment_training"],
    "other_ga": ["other g&a", "g&a lainnya", "other_ga"],
    "seed_investment": ["seed_investment", "seed investment"],
    "pre_money_valuation": ["pre_money_valuation", "pre-money valuation"],
    "exit_revenue_multiple_conservative": ["exit_revenue_multiple_conservative", "exit revenue multiple - conservative"],
    "exit_revenue_multiple_base": ["exit_revenue_multiple_base", "exit revenue multiple - base"],
    "exit_revenue_multiple_optimistic": ["exit_revenue_multiple_optimistic", "exit revenue multiple - optimistic"],
}

def clean_num(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        cleaned = re.sub(r'[Rp$€,%\s]', '', val)
        try:
            return float(cleaned)
        except ValueError:
            return 0.0
    return 0.0

def find_year_cols(sheet):
    year_cols = {}
    for row in range(1, 15):
        for col in range(1, 30):
            val = sheet.cell(row=row, column=col).value
            if val is not None:
                val_str = str(val).strip()
                match = re.match(r'^(20\d{2})$', val_str)
                if match:
                    y = int(match.group(1))
                    if y not in year_cols:
                        year_cols[y] = col
    if not year_cols:
        default_years = [2025, 2026, 2027, 2028, 2029]
        for idx, y in enumerate(default_years, start=2):
            year_cols[y] = idx
    return year_cols

def parse_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Locate main assumption sheet
    main_sheet = None
    target_sheet_names = ['02_assumptions', 'assumptions', 'asumsi']
    for name in wb.sheetnames:
        if any(t in name.lower() for t in target_sheet_names):
            main_sheet = wb[name]
            break
    if not main_sheet:
        main_sheet = wb.active

    year_cols = find_year_cols(main_sheet)
    sorted_years = sorted(year_cols.keys())
    result_by_year = {y: {} for y in sorted_years}

    # 1. Read main assumption sheet
    for row_idx, key in STANDARD_ROW_MAP.items():
        for y, col_idx in year_cols.items():
            raw_val = main_sheet.cell(row=row_idx, column=col_idx).value
            result_by_year[y][key] = clean_num(raw_val)

    # 2. Read 06_HR_Planning sheet if present
    hr_sheet = None
    for name in wb.sheetnames:
        if 'hr_planning' in name.lower() or 'hr' in name.lower() or 'payroll' in name.lower():
            hr_sheet = wb[name]
            break
    if hr_sheet:
        hr_year_cols = find_year_cols(hr_sheet)
        for row_idx, key in HR_ROW_MAP.items():
            for y in sorted_years:
                col_idx = hr_year_cols.get(y, year_cols.get(y))
                if col_idx:
                    raw_val = hr_sheet.cell(row=row_idx, column=col_idx).value
                    result_by_year[y][key] = clean_num(raw_val)

    # 3. Read 07_OPEX sheet if present
    opex_sheet = None
    for name in wb.sheetnames:
        if 'opex' in name.lower() or 'operating' in name.lower():
            opex_sheet = wb[name]
            break
    if opex_sheet:
        opex_year_cols = find_year_cols(opex_sheet)
        for row_idx, key in OPEX_ROW_MAP.items():
            for y in sorted_years:
                col_idx = opex_year_cols.get(y, year_cols.get(y))
                if col_idx:
                    raw_val = opex_sheet.cell(row=row_idx, column=col_idx).value
                    # Don't overwrite payroll_cost if already populated from HR Planning unless non-zero
                    val = clean_num(raw_val)
                    if key != 'payroll_cost' or val > 0:
                        result_by_year[y][key] = val

    # 4. Generic Mode scan across all sheets for missing keys
    for s_name in wb.sheetnames:
        sh = wb[s_name]
        sh_year_cols = find_year_cols(sh)
        for row_idx in range(1, sh.max_row + 1):
            label = str(sh.cell(row=row_idx, column=1).value or '').strip().lower()
            if not label:
                continue
            for key, aliases in KEY_ALIASES.items():
                if any(alias in label for alias in aliases):
                    for y in sorted_years:
                        col_idx = sh_year_cols.get(y, year_cols.get(y))
                        if col_idx:
                            if key not in result_by_year[y] or result_by_year[y][key] == 0:
                                raw_val = sh.cell(row=row_idx, column=col_idx).value
                                result_by_year[y][key] = clean_num(raw_val)

    # Normalize percentage fraction keys to whole percentage numbers for database & UI (e.g. 0.01 -> 1.0 %, 0.4 -> 40.0 %)
    for y in sorted_years:
        for k in FRAC_KEYS:
            if k in result_by_year[y]:
                val = result_by_year[y][k]
                if val is not None and 0 < val <= 1.0:
                    result_by_year[y][k] = round(val * 100.0, 4)

    return {
        "success": True,
        "sheet_name": main_sheet.title,
        "years": sorted_years,
        "assumptions": result_by_year
    }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "No file path provided"}))
        sys.exit(1)

    file_path = sys.argv[1]
    try:
        data = parse_excel(file_path)
        print(json.dumps(data))
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
        sys.exit(1)
