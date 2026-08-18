import sys
import json
import copy
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

def get_val(data, keys, default=None):
    for k in keys:
        if k in data and data[k] is not None:
            return data[k]
    return default

def copy_cell(src_cell, dst_cell, src_coord, dst_coord):
    """Copy formatting and translate formulas from src_cell to dst_cell."""
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)
        dst_cell.alignment = copy.copy(src_cell.alignment)

    val = src_cell.value
    if isinstance(val, str) and val.startswith('='):
        try:
            dst_cell.value = Translator(val, origin=src_coord).translate_formula(dst_coord)
        except Exception:
            dst_cell.value = val
    elif val is not None:
        dst_cell.value = val

def export_financial_model(json_data_path, template_path, output_path):
    with open(json_data_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)

    assumptions_by_year = payload.get('assumptions', {})
    company_name = payload.get('company_name', 'Smartcoop')
    currency = str(payload.get('currency', 'IDR')).upper()
    lang = str(payload.get('lang', 'en')).lower()

    # Currency rate and format configuration
    if currency == 'USD':
        rate = 1.0 / 17000.0
        num_format = '"$"#,##0'
    elif currency == 'EUR':
        rate = 1.0 / 20000.0
        num_format = '"€"#,##0'
    else:
        currency = 'IDR'
        rate = 1.0
        num_format = '"Rp "#,##0'

    # Determine sorted years from payload keys
    payload_years = []
    for k in assumptions_by_year.keys():
        try:
            payload_years.append(int(k))
        except (ValueError, TypeError):
            pass
    
    payload_years = sorted(list(set(payload_years)))
    if not payload_years:
        payload_years = [2025, 2026, 2027, 2028, 2029]

    wb = openpyxl.load_workbook(template_path, data_only=False)

    # 1. Update Cover Title & Subtitles according to Language & Currency
    if '01_Cover' in wb.sheetnames:
        sheet_cover = wb['01_Cover']
        if lang == 'id':
            sheet_cover['B3'] = f"Model Keuangan Pro-Forma v2.0 ({currency}) — {company_name}"
            sheet_cover['B4'] = f"Model Proyeksi Keuangan Koperasi 5-Tahun ({currency})"
        else:
            sheet_cover['B3'] = f"Financial Model v2.0 ({currency}) — {company_name}"
            sheet_cover['B4'] = f"5-Year Financial Projection Model ({currency})"

    model_sheets = [
        '02_Assumptions', '03_Customer_Growth', '04_Revenue_Engine', '05_COGS',
        '06_HR_Planning', '07_OPEX', '08_EBITDA', '09_Cash_Flow', '10_SaaS_Metrics',
        '11_Valuation'
    ]

    num_years = len(payload_years)

    # Label translations for Indonesian if lang == 'id'
    label_translations_id = {
        'Metric / Driver': 'Metrik / Asumsi Driver',
        '02. Financial Assumptions & Key Drivers': '02. Asumsi Keuangan & Driver Utama',
        '03. Customer Growth Model': '03. Model Pertumbuhan Koperasi Pelanggan',
        '04. Revenue Engine': '04. Mesin Pendapatan (Revenue)',
        '05. Cost of Revenue (COGS)': '05. Beban Pokok Pendapatan (COGS)',
        '06. HR Planning': '06. Perencanaan SDM (HR)',
        '07. Operating Expenses': '07. Beban Operasional (OPEX)',
        '08. EBITDA Summary': '08. Ringkasan EBITDA Summary',
        '09. Cash Flow Statement': '09. Laporan Arus Kas (Cash Flow)',
        '10. Key SaaS Metrics': '10. Metrik Utama SaaS',
        '11. Valuation & Exit Analysis': '11. Analisis Valuasi & Exit',
    }

    # Update headers and extend columns across sheets if necessary
    for sheet_name in model_sheets:
        if sheet_name not in wb.sheetnames:
            continue
        sheet = wb[sheet_name]

        # Translate section title at A1 and header at A3 if lang == 'id'
        if lang == 'id':
            if sheet.cell(1, 1).value in label_translations_id:
                sheet.cell(1, 1).value = label_translations_id[sheet.cell(1, 1).value]
            if sheet.cell(3, 1).value in label_translations_id:
                sheet.cell(3, 1).value = label_translations_id[sheet.cell(3, 1).value]

        # First extend columns if num_years > 5
        for i in range(num_years):
            col_idx = 2 + i
            if col_idx > 6:
                ref_col_letter = get_column_letter(6)
                target_col_letter = get_column_letter(col_idx)
                for r in range(1, sheet.max_row + 1):
                    src = sheet.cell(r, 6)
                    dst = sheet.cell(r, col_idx)
                    src_coord = f"{ref_col_letter}{r}"
                    dst_coord = f"{target_col_letter}{r}"
                    copy_cell(src, dst, src_coord, dst_coord)
        
        # Set exact year headers in Row 3
        for i, yr in enumerate(payload_years):
            col_idx = 2 + i
            sheet.cell(3, col_idx).value = yr
            sheet.cell(3, col_idx).number_format = '0'

    # Update 14_Dashboard headers if present
    if '14_Dashboard' in wb.sheetnames:
        sheet_dash = wb['14_Dashboard']
        if lang == 'id':
            sheet_dash['A3'] = f"Model Keuangan Smartcoop v2.0 - Hasil Utama ({currency})"
            sheet_dash['A11'] = "Metrik / Hasil Utama"
        else:
            sheet_dash['A3'] = f"Smartcoop Financial Model v2.0 - Key Outputs ({currency})"

        for i in range(num_years):
            col_idx = 2 + i
            if col_idx > 6:
                ref_col_letter = get_column_letter(6)
                target_col_letter = get_column_letter(col_idx)
                for r in range(11, 20):
                    src = sheet_dash.cell(r, 6)
                    dst = sheet_dash.cell(r, col_idx)
                    src_coord = f"{ref_col_letter}{r}"
                    dst_coord = f"{target_col_letter}{r}"
                    copy_cell(src, dst, src_coord, dst_coord)

        for i, yr in enumerate(payload_years):
            col_idx = 2 + i
            sheet_dash.cell(11, col_idx).value = yr

    sheet_a = wb['02_Assumptions']
    sheet_hr = wb['06_HR_Planning'] if '06_HR_Planning' in wb.sheetnames else None
    sheet_opex = wb['07_OPEX'] if '07_OPEX' in wb.sheetnames else None

    # Apply currency number formatting on monetary rows across sheets
    monetary_row_map = {
        '02_Assumptions': [13, 15, 16, 19, 22, 24, 26, 27, 30, 31, 32, 37, 38],
        '04_Revenue_Engine': [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15],
        '05_COGS': [5, 6, 7, 8, 9, 10, 11],
        '06_HR_Planning': [12, 13],
        '07_OPEX': [5, 6, 7, 8, 9, 10, 11, 12, 13],
        '08_EBITDA': [5, 6, 7, 9, 10],
        '09_Cash_Flow': [5, 6, 7, 8, 9],
        '10_SaaS_Metrics': [5, 6, 7, 11, 12, 14],
        '11_Valuation': [5, 6, 10, 11, 12, 13, 14],
        '13_Investor_Return': [5, 7, 9, 10],
        '14_Dashboard': [14, 15, 17, 19]
    }

    for sheet_name, rows in monetary_row_map.items():
        if sheet_name in wb.sheetnames:
            st = wb[sheet_name]
            for r in rows:
                for i in range(num_years):
                    col_idx = 2 + i
                    st.cell(r, col_idx).number_format = num_format

    # Populate assumption values for each year
    for i, yr in enumerate(payload_years):
        col_idx = 2 + i
        col_letter = get_column_letter(col_idx)
        prev_col_letter = get_column_letter(col_idx - 1) if i > 0 else None

        data = assumptions_by_year.get(str(yr)) or assumptions_by_year.get(yr) or {}

        # Row 6: Beginning Active Cooperatives
        if i == 0:
            beg_coops = get_val(data, ['beginning_cooperatives', 'initial_cooperatives', 'beginningCoops'])
            if beg_coops is not None:
                sheet_a.cell(6, col_idx).value = float(beg_coops)
        else:
            sheet_a.cell(6, col_idx).value = f"='03_Customer_Growth'!{prev_col_letter}8"

        # Row 7: New Cooperatives Acquired
        v = get_val(data, ['new_coops_acquired', 'newCoops'])
        if v is not None: sheet_a.cell(7, col_idx).value = float(v)

        # Row 8: Monthly Churn Rate (Percentage -> divide by 100)
        v = get_val(data, ['monthly_churn_rate', 'monthlyChurnRate'])
        if v is not None: sheet_a.cell(8, col_idx).value = float(v) / 100.0

        # Row 9: Average Members / Cooperative
        v = get_val(data, ['avg_members_per_coop', 'avgMembersPerCoop', 'avgMembers'])
        if v is not None: sheet_a.cell(9, col_idx).value = float(v)

        # Row 10: Subscription Paying % (Percentage -> divide by 100)
        v = get_val(data, ['subscription_paying_frac', 'subscriptionPayingFrac'])
        if v is not None: sheet_a.cell(10, col_idx).value = float(v) / 100.0

        # Row 13: Setup Fee / Paid Implementation Coop (Monetary -> multiply rate)
        v = get_val(data, ['setup_fee', 'setup_fee_per_coop', 'setupFee'])
        if v is not None: sheet_a.cell(13, col_idx).value = float(v) * rate

        # Row 14: Paid Implementation Coops
        v = get_val(data, ['paid_implementation_coops', 'paidImplementationCoops'])
        if v is not None: sheet_a.cell(14, col_idx).value = float(v)

        # Row 15: Monthly Subscription Fee (Monetary -> multiply rate)
        v = get_val(data, ['monthly_subscription_fee', 'saas_subscription_fee_per_coop', 'monthlySubscriptionFee'])
        if v is not None: sheet_a.cell(15, col_idx).value = float(v) * rate

        # Row 16: iOS Add-on Monthly Fee (Monetary -> multiply rate)
        v = get_val(data, ['ios_addon_monthly_fee', 'ios_addon_fee_per_coop', 'iosAddonMonthlyFee'])
        if v is not None: sheet_a.cell(16, col_idx).value = float(v) * rate

        # Row 17: iOS Adoption % (Percentage -> divide by 100)
        v = get_val(data, ['ios_adoption_frac', 'ios_adoption_rate', 'iosAdoptionRate'])
        if v is not None: sheet_a.cell(17, col_idx).value = float(v) / 100.0

        # Row 18: White Label Projects
        v = get_val(data, ['white_label_projects', 'white_label_projects_count', 'whiteLabelProjects'])
        if v is not None: sheet_a.cell(18, col_idx).value = float(v)

        # Row 19: White Label Fee / Project (Monetary -> multiply rate)
        v = get_val(data, ['white_label_fee_per_project', 'white_label_price_per_project', 'whiteLabelFeePerProject'])
        if v is not None: sheet_a.cell(19, col_idx).value = float(v) * rate

        # Row 20: PPOB Active Cooperatives % (Percentage -> divide by 100)
        v = get_val(data, ['ppob_active_coops_frac', 'ppob_adoption_rate', 'ppobAdoptionRate'])
        if v is not None: sheet_a.cell(20, col_idx).value = float(v) / 100.0

        # Row 21: PPOB Transactions / Active Coop / Month
        v = get_val(data, ['ppob_tx_per_coop_month', 'ppob_transactions_per_coop_month', 'ppobTxPerCoopMonth'])
        if v is not None: sheet_a.cell(21, col_idx).value = float(v)

        # Row 22: Average PPOB Fee / Transaction (Monetary -> multiply rate)
        v = get_val(data, ['avg_ppob_fee_per_tx', 'ppob_fee_per_transaction', 'avgPpobFeePerTx'])
        if v is not None: sheet_a.cell(22, col_idx).value = float(v) * rate

        # Row 23: Academy Participants % of Members (Percentage -> divide by 100)
        v = get_val(data, ['academy_participants_frac', 'academy_adoption_rate', 'academyAdoptionRate'])
        if v is not None: sheet_a.cell(23, col_idx).value = float(v) / 100.0

        # Row 24: Academy Average Price / Participant (Monetary -> multiply rate)
        v = get_val(data, ['academy_avg_price_per_participant', 'academy_price_per_participant', 'academyPricePerParticipant'])
        if v is not None: sheet_a.cell(24, col_idx).value = float(v) * rate

        # Row 25: Offline Trainings / Month
        v = get_val(data, ['offline_trainings_per_month', 'offlineTrainingsPerMonth'])
        if v is not None: sheet_a.cell(25, col_idx).value = float(v)

        # Row 26: Offline Training Fee / Cooperative (Monetary -> multiply rate)
        v = get_val(data, ['offline_training_fee_per_coop', 'offlineTrainingFeePerCoop'])
        if v is not None: sheet_a.cell(26, col_idx).value = float(v) * rate

        # Row 27: Enterprise / Banking / API Revenue (Monetary -> multiply rate)
        v = get_val(data, ['enterprise_api_revenue', 'enterprise_api_contracts_revenue', 'enterpriseApiRevenue'])
        if v is not None: sheet_a.cell(27, col_idx).value = float(v) * rate

        # Row 30: Cloud Cost / Active Coop / Month (Monetary -> multiply rate)
        v = get_val(data, ['cloud_cost_per_coop_month', 'cloudCostPerCoopMonth'])
        if v is not None: sheet_a.cell(30, col_idx).value = float(v) * rate

        # Row 31: Implementation Cost / Paid Implementation Coop (Monetary -> multiply rate)
        v = get_val(data, ['implementation_cost_per_coop', 'implementationCostPerCoop'])
        if v is not None: sheet_a.cell(31, col_idx).value = float(v) * rate

        # Row 32: Support Cost / Active Coop / Month (Monetary -> multiply rate)
        v = get_val(data, ['support_cost_per_coop_month', 'supportCostPerCoopMonth'])
        if v is not None: sheet_a.cell(32, col_idx).value = float(v) * rate

        # Row 33: Payment / API Variable Cost % of PPOB Revenue (Percentage -> divide by 100)
        v = get_val(data, ['payment_api_var_cost_frac', 'paymentApiVarCostFrac'])
        if v is not None: sheet_a.cell(33, col_idx).value = float(v) / 100.0

        # Row 34: Other Cost of Revenue % (Percentage -> divide by 100)
        v = get_val(data, ['other_cost_of_revenue_frac', 'otherCostOfRevenueFrac'])
        if v is not None: sheet_a.cell(34, col_idx).value = float(v) / 100.0

        # Row 37: Seed Investment (Monetary -> multiply rate)
        v = get_val(data, ['seed_investment', 'seedInvestment'])
        if v is not None: sheet_a.cell(37, col_idx).value = float(v) * rate

        # Row 38: Pre-Money Valuation (Monetary -> multiply rate)
        v = get_val(data, ['pre_money_valuation', 'preMoneyValuation'])
        if v is not None: sheet_a.cell(38, col_idx).value = float(v) * rate

        # Row 39: Exit Revenue Multiple - Conservative
        v = get_val(data, ['exit_revenue_multiple_conservative', 'exitRevenueMultipleConservative'])
        if v is not None: sheet_a.cell(39, col_idx).value = float(v)

        # Row 40: Exit Revenue Multiple - Base Case
        v = get_val(data, ['exit_revenue_multiple_base', 'exitRevenueMultipleBase'])
        if v is not None: sheet_a.cell(40, col_idx).value = float(v)

        # Row 41: Exit Revenue Multiple - Optimistic
        v = get_val(data, ['exit_revenue_multiple_conservative', 'exitRevenueMultipleOptimistic'])
        if v is not None: sheet_a.cell(41, col_idx).value = float(v)

        # ----------------------------------------------------
        # 06_HR_Planning Sheet
        # ----------------------------------------------------
        if sheet_hr is not None:
            v = get_val(data, ['hr_engineering_fte'])
            if v is not None: sheet_hr.cell(5, col_idx).value = float(v)

            v = get_val(data, ['hr_sales_fte'])
            if v is not None: sheet_hr.cell(6, col_idx).value = float(v)

            v = get_val(data, ['hr_marketing_fte'])
            if v is not None: sheet_hr.cell(7, col_idx).value = float(v)

            v = get_val(data, ['hr_support_fte'])
            if v is not None: sheet_hr.cell(8, col_idx).value = float(v)

            v = get_val(data, ['hr_finance_admin_fte'])
            if v is not None: sheet_hr.cell(9, col_idx).value = float(v)

            v = get_val(data, ['hr_management_fte'])
            if v is not None: sheet_hr.cell(10, col_idx).value = float(v)

            # Row 12: Average Salary (Monetary -> multiply rate)
            v = get_val(data, ['hr_avg_salary_monthly'])
            if v is not None: sheet_hr.cell(12, col_idx).value = float(v) * rate

        # ----------------------------------------------------
        # 07_OPEX Sheet
        # ----------------------------------------------------
        if sheet_opex is not None:
            # Row 5: Payroll Cost (Monetary -> multiply rate)
            v = get_val(data, ['payroll_cost', 'payrollCost'])
            if v is not None: sheet_opex.cell(5, col_idx).value = float(v) * rate

            # Row 6: Sales & Marketing Spend (Monetary -> multiply rate)
            v = get_val(data, ['sales_marketing_spend', 'salesMarketingSpend'])
            if v is not None: sheet_opex.cell(6, col_idx).value = float(v) * rate

            # Row 7: Office Utilities Internet (Monetary -> multiply rate)
            v = get_val(data, ['office_utilities_internet', 'officeUtilitiesInternet'])
            if v is not None: sheet_opex.cell(7, col_idx).value = float(v) * rate

            # Row 8: Software Tools Subscriptions (Monetary -> multiply rate)
            v = get_val(data, ['software_tools_subscriptions', 'softwareToolsSubscriptions'])
            if v is not None: sheet_opex.cell(8, col_idx).value = float(v) * rate

            # Row 9: Legal Accounting Compliance (Monetary -> multiply rate)
            v = get_val(data, ['legal_accounting_compliance', 'legalAccountingCompliance'])
            if v is not None: sheet_opex.cell(9, col_idx).value = float(v) * rate

            # Row 10: Travel Events (Monetary -> multiply rate)
            v = get_val(data, ['travel_events', 'travelEvents'])
            if v is not None: sheet_opex.cell(10, col_idx).value = float(v) * rate

            # Row 11: Recruitment Training (Monetary -> multiply rate)
            v = get_val(data, ['recruitment_training', 'recruitmentTraining'])
            if v is not None: sheet_opex.cell(11, col_idx).value = float(v) * rate

            # Row 12: Other GA (Monetary -> multiply rate)
            v = get_val(data, ['other_ga', 'otherGa'])
            if v is not None: sheet_opex.cell(12, col_idx).value = float(v) * rate

    wb.save(output_path)
    print(f"Successfully generated dynamic Excel model ({currency}/{lang}) at {output_path}")

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Usage: python export_excel.py <json_data_path> <template_path> <output_path>")
        sys.exit(1)
    
    export_financial_model(sys.argv[1], sys.argv[2], sys.argv[3])
